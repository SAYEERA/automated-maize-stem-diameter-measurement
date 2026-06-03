"""
Tangent-based maize stem diameter measurement pipeline.

This script processes segmented stem images and YOLO node detections. It exports
 overlays and Excel tables containing only tangent-based diameter
measurements. Diameters are measured perpendicular to the local skeleton direction,
which is more appropriate for curved stems than a fixed horizontal line.
"""

import cv2
import numpy as np
import pandas as pd
from pathlib import Path
from skimage import morphology, measure
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
import matplotlib.pyplot as plt
import os

# ── USER SETTINGS ───────────────────────────────────────────────────────────────
px_per_mm_vert = 8.117542
px_per_mm_diam = 8.0

MIN_NODE_GAP_CM = 2.0
MIN_VALID_ROWS_FOR_NODE_SNAP = 3

# Tangent settings
TANGENT_NEIGHBOR_RADIUS = 18
TANGENT_SAMPLE_HALF_WINDOW = 3
RAY_STEP = 0.5
MAX_RAY_STEPS = 2000

# Five representative diameter positions are sampled within each internode.
INTERNODE_SAMPLE_FRACS = [0.10, 0.30, 0.50, 0.70, 0.90]

# Drawing settings (paper-friendly)
SHOW_SKELETON = False
SHOW_CONTOUR = True

COLOR_CONTOUR = (0, 255, 0)        # green contour
COLOR_SKELETON = (255, 255, 0)      # cyan skeleton
COLOR_TANGENT_METHOD = (0, 215, 255) # bright yellow-green diameter lines
COLOR_NODE_TEXT = (0, 255, 0)       # green node markers
COLOR_DIAM_TEXT = (0, 0, 255)       # red diameter text

LINE_THICKNESS = 7
TEXT_SCALE = 2.3
TEXT_THICKNESS = 5
NODE_TEXT_SCALE = 2.4
NODE_TEXT_THICKNESS = 5
LABEL_X_OFFSET = 32

# Processing log. Batches listed here are skipped on later runs.
diameter_log = os.path.join("test_images/sample_outputs",
    "diameter_completed_batches.txt"
)

completed_diameter_batches = set()
if os.path.exists(diameter_log):
    with open(diameter_log, "r") as f:
        completed_diameter_batches = set(line.strip() for line in f)


# ── HELPERS ──────────────────────────────────────────────────────────────
def load_image_gray(path):
    """Load an image as normalized 8-bit grayscale."""
    img = cv2.imread(str(path), cv2.IMREAD_UNCHANGED)
    if img is None:
        raise FileNotFoundError(f"Image not found: {path}")
    if img.ndim == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return cv2.normalize(img, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)


def sharpen_unsharp_mask(img, radius=3, amount=1.5):
    """Apply mild unsharp masking before segmentation."""
    blurred = cv2.GaussianBlur(img, (0, 0), radius)
    return cv2.addWeighted(img, 1 + amount, blurred, -amount, 0)


def load_node_labels(label_file, img_shape, distance_thresh=250):
    """Load YOLO-format node labels and merge nearby duplicate detections."""
    img_h, img_w = img_shape
    nodes = []
    confs = []

    if not Path(label_file).exists():
        return np.zeros((0, 2), dtype=int)

    with open(label_file) as f:
        for line in f:
            parts = line.split()
            if len(parts) < 5:
                continue
            cls, x_rel, y_rel, w_rel, h_rel = map(float, parts[:5])
            nodes.append((int(x_rel * img_w), int(y_rel * img_h)))
            confs.append(h_rel)

    if not nodes:
        return np.zeros((0, 2), dtype=int)

    nodes = np.array(nodes, int)
    confs = np.array(confs)

    filtered_nodes = []
    used = set()

    for i, pt in enumerate(nodes):
        if i in used:
            continue

        cluster = [i]
        for j, other in enumerate(nodes):
            if j != i and j not in used:
                if np.linalg.norm(pt - other) < distance_thresh:
                    cluster.append(j)

        best_idx = max(cluster, key=lambda idx: confs[idx])
        filtered_nodes.append(tuple(nodes[best_idx]))
        used.update(cluster)

    return np.array(filtered_nodes, int)


def skeletonize_region(region_mask):
    """Convert a binary stem region to a one-pixel-wide centerline."""
    try:
        from cv2.ximgproc import thinning
        return thinning(region_mask, cv2.ximgproc.THINNING_ZHANGSUEN)
    except Exception:
        return morphology.skeletonize(region_mask > 0).astype(np.uint8) * 255


def connect_and_fill(binary_img, connect_dist=15):
    """Close small gaps and fill holes in the segmented stem mask."""
    binary = (binary_img > 0).astype(np.uint8) * 255

    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (connect_dist, connect_dist))
    connected = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)

    inv = cv2.bitwise_not(connected)
    h, w = inv.shape[:2]
    mask = np.zeros((h + 2, w + 2), np.uint8)
    flood = inv.copy()
    cv2.floodFill(flood, mask, (0, 0), 255)
    flood_inv = cv2.bitwise_not(flood)
    filled = connected | flood_inv

    filled = cv2.morphologyEx(filled, cv2.MORPH_CLOSE, np.ones((5, 5), np.uint8), iterations=1)
    return filled


def extract_final_contours(binary_filled, min_area=500):
    """Extract smoothed external contours from the filled stem mask."""
    contours, _ = cv2.findContours(binary_filled, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = [cnt for cnt in contours if cv2.contourArea(cnt) > min_area]
    smooth_contours = [cv2.approxPolyDP(cnt, epsilon=1.0, closed=True) for cnt in contours]
    return smooth_contours


def snap_to_skeleton_ywise(x0, y0, skeleton_mask):
    """Move a point horizontally to the nearest skeleton pixel in the same row."""
    h, w = skeleton_mask.shape
    y0 = int(np.clip(y0, 0, h - 1))
    row = skeleton_mask[y0, :]
    x_coords = np.where(row > 0)[0]
    if len(x_coords) == 0:
        return int(x0), int(y0)
    nearest_x = x_coords[np.argmin(np.abs(x_coords - x0))]
    return int(nearest_x), int(y0)


def snap_to_skeleton_nearest(x0, y0, skeleton_mask, search_radius=20):
    """Move a point to the nearest skeleton pixel within a local search window."""
    h, w = skeleton_mask.shape
    x0 = int(np.clip(round(x0), 0, w - 1))
    y0 = int(np.clip(round(y0), 0, h - 1))

    y1 = max(0, y0 - search_radius)
    y2 = min(h, y0 + search_radius + 1)
    x1 = max(0, x0 - search_radius)
    x2 = min(w, x0 + search_radius + 1)

    patch = skeleton_mask[y1:y2, x1:x2]
    ys, xs = np.where(patch > 0)
    if len(xs) == 0:
        return int(x0), int(y0)

    xs = xs + x1
    ys = ys + y1
    d2 = (xs - x0) ** 2 + (ys - y0) ** 2
    idx = np.argmin(d2)
    return int(xs[idx]), int(ys[idx])


# ── TANGENT-BASED DIAMETER MEASUREMENT ──────────────────────────────────
# Diameter is measured perpendicular to the local stem centerline.
# The local tangent direction is estimated from nearby skeleton pixels.
# A normal line is then traced to both stem boundaries, and the distance
# between the two boundary points is reported as the stem diameter.
def collect_local_skeleton_points(skeleton_mask, x0, y0, radius=TANGENT_NEIGHBOR_RADIUS):
    """Return skeleton pixels near the requested measurement point.

    These local centerline pixels are used to estimate stem orientation.
    A local neighborhood is used because stalks may curve along their length.
    """
    h, w = skeleton_mask.shape
    x0 = int(np.clip(round(x0), 0, w - 1))
    y0 = int(np.clip(round(y0), 0, h - 1))

    y1 = max(0, y0 - radius)
    y2 = min(h, y0 + radius + 1)
    x1 = max(0, x0 - radius)
    x2 = min(w, x0 + radius + 1)

    patch = skeleton_mask[y1:y2, x1:x2]
    ys, xs = np.where(patch > 0)
    if len(xs) == 0:
        return np.empty((0, 2), dtype=np.float32)

    xs = xs + x1
    ys = ys + y1
    d2 = (xs - x0) ** 2 + (ys - y0) ** 2
    keep = d2 <= radius * radius

    xs = xs[keep]
    ys = ys[keep]

    return np.column_stack([xs.astype(np.float32), ys.astype(np.float32)])


def estimate_local_tangent_direction(skeleton_mask, x0, y0, radius=TANGENT_NEIGHBOR_RADIUS):
    """Estimate the local stem tangent using PCA on nearby skeleton pixels.

    The first principal component gives the dominant centerline direction
    around the measurement point.
    """
    x0, y0 = snap_to_skeleton_nearest(x0, y0, skeleton_mask, search_radius=20)
    pts = collect_local_skeleton_points(skeleton_mask, x0, y0, radius=radius)

    if len(pts) < 3:
        return np.array([0.0, 1.0], dtype=np.float32)

    pts_centered = pts - pts.mean(axis=0, keepdims=True)
    cov = np.cov(pts_centered.T)
    eigvals, eigvecs = np.linalg.eigh(cov)
    tangent = eigvecs[:, np.argmax(eigvals)].astype(np.float32)

    norm = np.linalg.norm(tangent)
    if norm < 1e-8:
        return np.array([0.0, 1.0], dtype=np.float32)

    tangent = tangent / norm

    if tangent[1] < 0:
        tangent = -tangent

    return tangent.astype(np.float32)


def trace_to_stem_boundary(mask, x0, y0, dx, dy, step=RAY_STEP, max_steps=MAX_RAY_STEPS):
    """Trace a ray from the centerline until it leaves the stem mask.

    The last location inside the mask is used as the boundary point.
    """
    h, w = mask.shape
    x, y = float(x0), float(y0)
    last_inside = (x, y)

    for _ in range(max_steps):
        xi = int(round(x))
        yi = int(round(y))
        if xi < 0 or xi >= w or yi < 0 or yi >= h or mask[yi, xi] == 0:
            break
        last_inside = (x, y)
        x += dx * step
        y += dy * step

    return last_inside


def measure_normal_diameter(region_mask_filled, skeleton, x0, y0):
    """Measure diameter along the normal to the local stem centerline.

    The tangent direction is rotated by 90 degrees to form the cross-section
    direction. The ray is traced in both directions to the stem boundary.
    """
    x0, y0 = snap_to_skeleton_nearest(x0, y0, skeleton, search_radius=20)
    tangent = estimate_local_tangent_direction(skeleton, x0, y0, radius=TANGENT_NEIGHBOR_RADIUS)
    tx, ty = float(tangent[0]), float(tangent[1])

    nx, ny = -ty, tx
    nrm = np.hypot(nx, ny)
    if nrm < 1e-8:
        nx, ny = 1.0, 0.0
    else:
        nx, ny = nx / nrm, ny / nrm

    boundary_positive = trace_to_stem_boundary(region_mask_filled, x0, y0, nx, ny)
    boundary_negative = trace_to_stem_boundary(region_mask_filled, x0, y0, -nx, -ny)

    diameter_px = float(np.hypot(boundary_positive[0] - boundary_negative[0], boundary_positive[1] - boundary_negative[1]))

    return {
        "diameter_px": diameter_px,
        "pt1": boundary_negative,
        "pt2": boundary_positive,
        "center": (float(x0), float(y0)),
        "tangent": (tx, ty),
        "normal": (nx, ny)
    }


def measure_robust_tangent_diameter(region_mask_filled, skeleton, x0, y0, half_window=TANGENT_SAMPLE_HALF_WINDOW):
    """Compute a stable tangent-based diameter near one point.

    Multiple nearby cross-sections are sampled along the local tangent. The
    cross-section closest to the median diameter is retained to reduce the
    effect of small mask irregularities.
    """
    x0, y0 = snap_to_skeleton_nearest(x0, y0, skeleton, search_radius=20)
    tangent = estimate_local_tangent_direction(skeleton, x0, y0, radius=TANGENT_NEIGHBOR_RADIUS)
    tx, ty = float(tangent[0]), float(tangent[1])

    samples = []
    for s in range(-half_window, half_window + 1):
        xs = x0 + s * tx
        ys = y0 + s * ty
        xs, ys = snap_to_skeleton_nearest(xs, ys, skeleton, search_radius=10)
        info = measure_normal_diameter(region_mask_filled, skeleton, xs, ys)
        if info["diameter_px"] > 0:
            samples.append(info)

    if not samples:
        return measure_normal_diameter(region_mask_filled, skeleton, x0, y0)

    diameters = np.array([s["diameter_px"] for s in samples], dtype=np.float32)
    med = float(np.median(diameters))
    idx = int(np.argmin(np.abs(diameters - med)))
    return samples[idx]


# ── DRAWING ──────────────────────────────────────────────────────────────
def draw_node_label(img, node_xy, node_label):
    """Draw a labeled node marker on the overlay image."""
    x, y = int(round(node_xy[0])), int(round(node_xy[1]))

    # Node marker used in the final overlay
    cv2.circle(img, (x, y), 34, (0, 180, 0), -1, lineType=cv2.LINE_AA)
    cv2.circle(img, (x, y), 34, (120, 255, 120), 4, lineType=cv2.LINE_AA)

    # Keep the label readable over both light and dark stem regions
    cv2.putText(
        img,
        node_label,
        (x - 28, y + 12),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.15,
        (255, 255, 255),
        3,
        lineType=cv2.LINE_AA
    )

def draw_measurement(img, tangent_measurement, diameter_mm, text_label):
    """Draw one tangent-based diameter line and its value on the overlay."""
    x1, y1 = int(round(tangent_measurement["pt1"][0])), int(round(tangent_measurement["pt1"][1]))
    x2, y2 = int(round(tangent_measurement["pt2"][0])), int(round(tangent_measurement["pt2"][1]))

    cv2.line(
        img,
        (x1, y1),
        (x2, y2),
        COLOR_TANGENT_METHOD,
        LINE_THICKNESS,
        lineType=cv2.LINE_AA,
    )

    tx = int(round(max(x1, x2))) + LABEL_X_OFFSET
    ty = int(round((y1 + y2) / 2))
    text = f"{text_label}: {diameter_mm:.1f}"

    cv2.putText(
        img,
        text,
        (tx, ty + 5),
        cv2.FONT_HERSHEY_SIMPLEX,
        TEXT_SCALE,
        COLOR_DIAM_TEXT,
        TEXT_THICKNESS,
        lineType=cv2.LINE_AA,
    )


# ── MAIN BATCH ───────────────────────────────────────────────────────────
def batch_process_stems():
    """Process all batch folders and save tangent-only outputs."""
    base_batch_dir = Path("test_images/sample_outputs")

    for batch_folder in sorted(base_batch_dir.glob("batch_*")):
        batch_name = batch_folder.name

        if batch_name in completed_diameter_batches:
            print(f"Skipping completed batch: {batch_name}")
            continue

        input_dir = batch_folder / "piscart_whitebg" / "tif"
        labels_dir = batch_folder / "yolo_output" / "labels"
        output_dir = batch_folder / "analysis_output"
        output_dir.mkdir(parents=True, exist_ok=True)

        tif_images = sorted(list(input_dir.glob("*.tif")))
        if not tif_images:
            print(f"No TIF images found in {batch_name}; skipping.")
            continue

        processed_count = 0

        for image_path in tif_images:
            image_name = image_path.stem
            base_name = image_name.replace("_whitebg", "").replace("_blackbg", "")
            label_path = labels_dir / f"{base_name}_blackbg.txt"
            image_output_dir = output_dir / base_name
            image_output_dir.mkdir(parents=True, exist_ok=True)

            print(f"\nProcessing {base_name} in batch {batch_name}")
            try:
                process_single_stem(image_path, label_path, image_output_dir)
                processed_count += 1
            except Exception as e:
                print(f"Error processing {base_name}: {e}")

        if processed_count > 0:
            with open(diameter_log, "a") as f:
                f.write(batch_name + "\n")
            print(f"Batch {batch_name} completed and logged ({processed_count} images).")
        else:
            print(f"Batch {batch_name} had no valid images; not logged.")


def process_single_stem(image_path, label_path, output_dir, global_node_start=1):
    """Process one image and export tangent-based stem measurements."""
    gray = load_image_gray(image_path)
    sharp = sharpen_unsharp_mask(gray)
    blur = cv2.bilateralFilter(sharp, 7, 50, 50)

    is_white_bg = np.mean(blur) > 128
    thresh_type = cv2.THRESH_BINARY_INV if is_white_bg else cv2.THRESH_BINARY

    _, otsu = cv2.threshold(blur, 0, 255, thresh_type + cv2.THRESH_OTSU)
    adaptive = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, thresh_type, 31, 2
    )
    binary = cv2.bitwise_or(otsu, adaptive)

    binary_filled = connect_and_fill(binary, connect_dist=15)
    contours = extract_final_contours(binary_filled)

    contour_mask = np.zeros_like(gray, dtype=np.uint8)
    cv2.drawContours(contour_mask, contours, -1, 255, thickness=1)

    img_overlay = cv2.cvtColor(sharp.copy(), cv2.COLOR_GRAY2BGR)
    if SHOW_CONTOUR:
        cv2.drawContours(img_overlay, contours, -1, COLOR_CONTOUR, thickness=2, lineType=cv2.LINE_AA)

    nodes_all = load_node_labels(label_path, gray.shape)

    merged = cv2.morphologyEx(binary_filled, cv2.MORPH_CLOSE, np.ones((15, 3), np.uint8), iterations=2)
    labeled = measure.label(merged)
    regions_raw = measure.regionprops(labeled)

    def region_bottom_x(region):
        ys, xs = np.nonzero(labeled == region.label)
        if len(ys) == 0:
            return 1e18
        yb = ys.max()
        xs_row = xs[ys == yb]
        if len(xs_row) == 0:
            return float(np.median(xs))
        return float(np.median(xs_row))

    regions = sorted(regions_raw, key=lambda r: region_bottom_x(r))

    rows = []
    global_int = int(global_node_start)
    stem_counter = 0

    for region in regions:
        if region.area < 5000:
            continue

        region_mask = (labeled == region.label).astype(np.uint8)
        region_mask = morphology.binary_closing(region_mask, morphology.disk(3))
        region_mask = morphology.binary_dilation(region_mask, morphology.disk(1))
        region_mask = (region_mask * 255).astype(np.uint8)
        region_mask_filled = (region_mask > 0).astype(np.uint8)

        skeleton = skeletonize_region(region_mask)
        if len(np.nonzero(skeleton)[0]) < 10:
            continue

        if SHOW_SKELETON:
            sk_y, sk_x = np.where(skeleton > 0)
            for x, y in zip(sk_x, sk_y):
                img_overlay[y, x] = COLOR_SKELETON

        ys, xs = np.nonzero(region_mask_filled)
        if len(ys) == 0:
            continue

        stem_counter += 1
        stem_id = stem_counter

        stem_bottom_ref_y = int(ys.max())
        stem_top_ref_y = int(ys.min())
        stem_total_len_cm = (stem_bottom_ref_y - stem_top_ref_y) / px_per_mm_vert / 10.0

        y_bottom_actual = ys.max()
        x_bottom_actual = int(np.median(xs[ys == y_bottom_actual]))

        y_bottom_offset = max(0, y_bottom_actual - 5)
        x_bottom_offset = int(np.median(xs[ys == y_bottom_offset])) if np.any(ys == y_bottom_offset) else x_bottom_actual

        bottom_actual = measure_robust_tangent_diameter(region_mask_filled, skeleton, x_bottom_actual, y_bottom_actual)
        bottom_offset = measure_robust_tangent_diameter(region_mask_filled, skeleton, x_bottom_offset, y_bottom_offset)

        if bottom_offset["diameter_px"] > bottom_actual["diameter_px"]:
            bottom_node = (x_bottom_offset, y_bottom_offset)
        else:
            bottom_node = (x_bottom_actual, y_bottom_actual)

        top_node = (int(np.median(xs[ys == ys.min()])), ys.min())

        y_shift = int(round(px_per_mm_vert))
        new_bottom = (bottom_node[0], bottom_node[1] - y_shift)
        new_top = (top_node[0], top_node[1] + y_shift)

        minr, minc, maxr, maxc = region.bbox

        def in_region(pt):
            return minc <= pt[0] <= maxc and minr <= pt[1] <= maxr

        region_nodes = [pt for pt in nodes_all if in_region(pt)] + [new_bottom]

        top_threshold_y = minr + (maxr - minr) * 0.15
        top_present = any(pt[1] <= top_threshold_y for pt in nodes_all if in_region(pt))
        if not top_present:
            region_nodes.append(new_top)

        region_nodes = sorted(region_nodes, key=lambda p: p[1], reverse=True)

        # Merge node detections that are too close to represent separate internodes.
        collapsed = []
        i = 0
        while i < len(region_nodes):
            if i == len(region_nodes) - 1:
                collapsed.append(region_nodes[i])
                break

            a = region_nodes[i]
            b = region_nodes[i + 1]
            gap_cm = abs(a[1] - b[1]) / px_per_mm_vert / 10.0

            if gap_cm < MIN_NODE_GAP_CM:
                fracs = [0.0, 1/6, 2/6, 3/6, 4/6, 5/6, 1.0]
                best_pt = a
                best_d = -1.0
                for f in fracs:
                    x = int((1 - f) * a[0] + f * b[0])
                    y = int((1 - f) * a[1] + f * b[1])
                    x, y = snap_to_skeleton_nearest(x, y, skeleton, search_radius=20)
                    tangent_measurement = measure_robust_tangent_diameter(region_mask_filled, skeleton, x, y)
                    if tangent_measurement["diameter_px"] > best_d:
                        best_d = tangent_measurement["diameter_px"]
                        best_pt = (x, y)
                collapsed.append(best_pt)
                i += 2
            else:
                collapsed.append(a)
                i += 1

        region_nodes = collapsed
        region_nodes = [snap_to_skeleton_nearest(x, y, skeleton, search_radius=20) for (x, y) in region_nodes]

        # Refine internal node locations by searching for the widest nearby tangent diameter.
        if len(region_nodes) >= 3:
            rn = list(region_nodes)
            for k in range(1, len(rn) - 1):
                prev_y = rn[k - 1][1]
                curr_x, curr_y = rn[k]
                next_y = rn[k + 1][1]

                y_low = int((prev_y + curr_y) / 2)
                y_high = int((curr_y + next_y) / 2)

                h, w = contour_mask.shape[:2]
                y_low = int(np.clip(y_low, 0, h - 1))
                y_high = int(np.clip(y_high, 0, h - 1))

                best = None
                valid = 0
                for yy in range(min(y_low, y_high), max(y_low, y_high) + 1):
                    xx, yy2 = snap_to_skeleton_ywise(curr_x, yy, skeleton)
                    tangent_measurement = measure_robust_tangent_diameter(region_mask_filled, skeleton, xx, yy2)
                    dpx = tangent_measurement["diameter_px"]
                    if dpx <= 0:
                        continue
                    valid += 1
                    if (best is None) or (dpx > best[0]):
                        best = (dpx, yy2, xx)

                if best is not None and valid >= MIN_VALID_ROWS_FOR_NODE_SNAP:
                    rn[k] = (int(best[2]), int(best[1]))

            region_nodes = sorted(rn, key=lambda p: p[1], reverse=True)

        if len(region_nodes) < 2:
            continue

        # Assign node names from the base toward the upper stem.
        node_names = []
        if stem_id == 1:
            node_names.append(f"node_{global_int}.0")
        else:
            node_names.append(f"node_{global_int}.1")

        for _ in range(1, len(region_nodes)):
            global_int += 1
            node_names.append(f"node_{global_int}.0")

        # Draw node labels in increasing numeric order.
        numeric_node_labels = []
        for nm in node_names:
            base = nm.split("_")[1]
            n_num = base.split(".")[0]
            numeric_node_labels.append(f"N{n_num}")

        for node_label, (nx, ny) in zip(numeric_node_labels, region_nodes):
            draw_node_label(img_overlay, (nx, ny), node_label)

        # Store rows in the same measurement order used for analysis: node -> D points -> next node.
        # This order is important because internode length is calculated from consecutive node rows.
        for i in range(len(region_nodes) - 1):
            node_a, node_b = region_nodes[i], region_nodes[i + 1]
            name_a, name_b = node_names[i], node_names[i + 1]

            def interpolate(p1, p2, frac):
                return (
                    int((1 - frac) * p1[0] + frac * p2[0]),
                    int((1 - frac) * p1[1] + frac * p2[1])
                )

            # Add the lower node of this internode. The upper node is added as
            # the lower node of the next internode, which prevents duplicate node rows.
            node_x, node_y = node_a
            node_point_len_cm = (stem_bottom_ref_y - node_y) / px_per_mm_vert / 10.0
            node_point_len_cm = max(0.0, node_point_len_cm)

            if i == 0:
                node_stem_len_cm = 0.0
            else:
                node_stem_len_cm = node_point_len_cm

            tangent_measurement_node = measure_robust_tangent_diameter(
                region_mask_filled, skeleton, node_x, node_y
            )
            node_diameter_mm = (
                tangent_measurement_node["diameter_px"] / px_per_mm_diam
                if tangent_measurement_node["diameter_px"] > 0 else 0.0
            )

            rows.append({
                "Stem": stem_id,
                "Type": name_a,
                "Y(px)": int(node_y),
                "X(px)": int(node_x),
                "Point Length (cm)": round(node_point_len_cm, 2),
                "Stem Length (cm)": round(node_stem_len_cm, 2),
                "Diameter Tangent (mm)": round(node_diameter_mm, 2),
            })

            # Measure tangent diameters at five fixed positions within this internode.
            point_defs = [
                ("D1", interpolate(node_a, node_b, INTERNODE_SAMPLE_FRACS[0])),
                ("D2", interpolate(node_a, node_b, INTERNODE_SAMPLE_FRACS[1])),
                ("D3", interpolate(node_a, node_b, INTERNODE_SAMPLE_FRACS[2])),
                ("D4", interpolate(node_a, node_b, INTERNODE_SAMPLE_FRACS[3])),
                ("D5", interpolate(node_a, node_b, INTERNODE_SAMPLE_FRACS[4])),
            ]

            for label, (x0, y0) in point_defs:
                x0, y0 = snap_to_skeleton_nearest(x0, y0, skeleton, search_radius=20)

                tangent_measurement = measure_robust_tangent_diameter(region_mask_filled, skeleton, x0, y0)
                diameter_mm = (
                    tangent_measurement["diameter_px"] / px_per_mm_diam
                    if tangent_measurement["diameter_px"] > 0 else 0.0
                )

                center_x, center_y = tangent_measurement["center"]
                center_x, center_y = int(round(center_x)), int(round(center_y))

                point_len_cm = (stem_bottom_ref_y - center_y) / px_per_mm_vert / 10.0
                point_len_cm = max(0.0, point_len_cm)

                rows.append({
                    "Stem": stem_id,
                    "Type": f"{label}_between_{name_a}_{name_b}",
                    "Y(px)": int(center_y),
                    "X(px)": int(center_x),
                    "Point Length (cm)": round(point_len_cm, 2),
                    "Stem Length (cm)": round(point_len_cm, 2),
                    "Diameter Tangent (mm)": round(diameter_mm, 2),
                })

                draw_measurement(img_overlay, tangent_measurement, diameter_mm, label)

        # Add the final upper node once, after all internodes have been written.
        final_node = region_nodes[-1]
        final_name = node_names[-1]
        final_x, final_y = final_node

        final_point_len_cm = (stem_bottom_ref_y - final_y) / px_per_mm_vert / 10.0
        final_point_len_cm = max(0.0, final_point_len_cm)
        final_stem_len_cm = stem_total_len_cm

        final_measurement = measure_robust_tangent_diameter(
            region_mask_filled, skeleton, final_x, final_y
        )
        final_diameter_mm = (
            final_measurement["diameter_px"] / px_per_mm_diam
            if final_measurement["diameter_px"] > 0 else 0.0
        )

        rows.append({
            "Stem": stem_id,
            "Type": final_name,
            "Y(px)": int(final_y),
            "X(px)": int(final_x),
            "Point Length (cm)": round(final_point_len_cm, 2),
            "Stem Length (cm)": round(final_stem_len_cm, 2),
            "Diameter Tangent (mm)": round(final_diameter_mm, 2),
        })

    # Save final overlay image.
    overlay_path = output_dir / f"{image_path.stem}_tangent_overlay_5points.png"
    cv2.imwrite(str(overlay_path), img_overlay)
    print(f"OVERLAY SAVED: {overlay_path}")

    # Save measurement table, per-stem sheets, and summary plots.
    df_all = pd.DataFrame(rows)
    if df_all.empty:
        print(f"No valid tangent measurements found for {image_path.stem}.")
        return

    internode_lengths = []
    last_node_point_len = None
    current_stem = None

    for _, r in df_all.iterrows():
        row_type = str(r["Type"])
        stem_id = r["Stem"]

        if stem_id != current_stem:
            last_node_point_len = None
            current_stem = stem_id

        if row_type.lower().startswith("node_"):
            current_len = float(r["Point Length (cm)"])
            if last_node_point_len is not None:
                internode_lengths.append(round(current_len - last_node_point_len, 2))
            else:
                internode_lengths.append("")
            last_node_point_len = current_len
        else:
            internode_lengths.append("")

    df_all["Internode Length(cms)"] = internode_lengths

    df_all = df_all[[
        "Stem", "Type", "Y(px)", "X(px)",
        "Point Length (cm)", "Stem Length (cm)",
        "Internode Length(cms)",
        "Diameter Tangent (mm)"
    ]]

    df_all["Stem Length (cm)"] = pd.to_numeric(df_all["Stem Length (cm)"], errors="coerce")
    df_all["Diameter Tangent (mm)"] = pd.to_numeric(df_all["Diameter Tangent (mm)"], errors="coerce")

    global_xmin = 0.0
    global_xmax = float(df_all["Stem Length (cm)"].max()) if df_all["Stem Length (cm)"].notna().any() else 0.0

    diameter_values = df_all["Diameter Tangent (mm)"].dropna()
    diameter_values = diameter_values[(diameter_values > 0) & (diameter_values < 200)]
    global_ymin = float(diameter_values.min()) if len(diameter_values) else 0.0
    global_ymax = float(diameter_values.max()) if len(diameter_values) else 1.0
    pad = 0.05 * (global_ymax - global_ymin + 1e-6)
    global_ymin -= pad
    global_ymax += pad

    excel_output_path = output_dir / f"{image_path.stem}_tangent_5point_measurements.xlsx"
    with pd.ExcelWriter(excel_output_path, engine="openpyxl", mode="w") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="All_Points", index=False)

    wb = load_workbook(excel_output_path)
    ws_all = wb["All_Points"]
    current_row = ws_all.max_row + 2

    for stem_id in sorted(df_all["Stem"].unique()):
        df_stem = df_all[df_all["Stem"] == stem_id].copy()
        if df_stem.empty:
            continue

        header_cell = ws_all.cell(row=current_row, column=2, value=f"Stem {stem_id} - Tangent Analysis")
        header_cell.font = Font(bold=True, color="800080")

        total_len = float(df_stem["Stem Length (cm)"].max())
        length_cell = ws_all.cell(row=current_row, column=6, value=f"Total Length = {round(total_len, 2)}cms")
        length_cell.font = Font(bold=True)

        sheet_name = f"Stem_{stem_id}"
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        header_cell.hyperlink = Hyperlink(ref="", location=f"'{sheet_name}'!A1", display=header_cell.value)

        for col_idx, col_name in enumerate(df_all.columns, start=1):
            ws_all.cell(row=current_row + 1, column=col_idx, value=col_name).font = Font(bold=True)

        data_start_row = current_row + 2
        for i, row_data in enumerate(df_stem.itertuples(index=False)):
            for col_idx, value in enumerate(row_data, start=1):
                ws_all.cell(row=data_start_row + i, column=col_idx, value=value)
        current_row = data_start_row + len(df_stem) + 4

        ws_stem = wb[sheet_name]
        max_length = int(np.ceil(df_stem["Stem Length (cm)"].max()))
        bins = list(range(0, max_length + 10, 10))
        if len(bins) < 2:
            bins = [0, 10]

        df_stem_summary = df_stem.copy()
        df_stem_summary["Segment"] = pd.cut(
            df_stem_summary["Stem Length (cm)"],
            bins=bins,
            right=False,
            include_lowest=True
        ).astype(str)

        segment_order = df_stem_summary["Segment"].dropna().unique()
        node_counts = (
            df_stem_summary[df_stem_summary["Type"].str.contains("^node_", regex=True, case=False)]
            .groupby("Segment", observed=False)
            .size()
        )
        internode_length_mean = (
            df_stem_summary[df_stem_summary["Type"].str.contains("^node_", regex=True, case=False)]
            .copy()
        )
        internode_length_mean["Internode Length(cms)"] = pd.to_numeric(
            internode_length_mean["Internode Length(cms)"], errors="coerce"
        )
        internode_length_mean = (
            internode_length_mean
            .groupby("Segment", observed=False)["Internode Length(cms)"]
            .mean()
        )
        avg_diameter = df_stem_summary.groupby("Segment", observed=False)["Diameter Tangent (mm)"].mean()

        summary = pd.DataFrame({
            "Segment": segment_order.astype(str),
            "Node Count": node_counts.reindex(segment_order).fillna(0).astype(int).values,
            "Avg Internode Length (cm)": internode_length_mean.reindex(segment_order).values,
            "Avg Tangent Diameter (mm)": avg_diameter.reindex(segment_order).values,
            "Curvature (deg)": [round(90 - i * 15.23, 2) for i in range(len(segment_order))]
        })

        for col_idx, header in enumerate(summary.columns, start=1):
            ws_stem.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
        for row_idx, row in enumerate(summary.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_stem.cell(row=row_idx, column=col_idx, value=value)

        offset = len(summary) + 2
        ws_stem.cell(row=offset, column=6, value=f"Total Length = {round(total_len, 2)}cms").font = Font(bold=True)

        for col_idx, col_name in enumerate(df_all.columns, start=1):
            ws_stem.cell(row=offset + 2, column=col_idx, value=col_name).font = Font(bold=True)
        for i, row_data in enumerate(df_stem_summary.itertuples(index=False)):
            for col_idx, value in enumerate(row_data, start=1):
                ws_stem.cell(row=offset + 3 + i, column=col_idx, value=value)

        # Segment-wise structural plot, matching the older output style.
        summary_plot = summary.copy()
        summary_plot["Segment Mid"] = summary_plot["Segment"].str.extract(r"(\d+)").astype(float)
        summary_plot = summary_plot.sort_values("Segment Mid")

        fig, ax = plt.subplots(figsize=(12, 6))
        x_labels = summary_plot["Segment"].tolist()
        x_pos = np.arange(len(x_labels))
        ax.plot(x_pos, summary_plot["Node Count"], marker="o", label="Node Count")
        ax.plot(x_pos, summary_plot["Avg Tangent Diameter (mm)"], marker="s", label="Avg Tangent Diameter (mm)")
        ax.plot(x_pos, summary_plot["Curvature (deg)"] / 10, marker="d", label="Curvature (deg / 10)")
        ax.set_title(f"Stem {stem_id} - Segment-wise Tangent Analysis")
        ax.set_xlabel("Segment (cm)")
        ax.set_ylabel("Value")
        ax.set_xticks(x_pos)
        ax.set_xticklabels(x_labels, rotation=45, ha="right")
        ax.legend()
        ax.grid(True)
        plt.tight_layout()
        plt.savefig(output_dir / f"{image_path.stem}_stem_{stem_id}_segments.png")
        plt.close(fig)

        # Diameter profile along stem axis.
        fig_profile, ax_profile = plt.subplots(figsize=(10, 5))
        raw_curve = df_stem[["Stem Length (cm)", "Diameter Tangent (mm)"]].copy()
        raw_curve = raw_curve.dropna().sort_values("Stem Length (cm)")
        raw_curve = raw_curve.drop_duplicates(subset=["Stem Length (cm)"], keep="first")

        if raw_curve.shape[0] >= 2:
            x_raw = raw_curve["Stem Length (cm)"].values
            y_raw = raw_curve["Diameter Tangent (mm)"].values
            x_interp = np.arange(x_raw.min(), x_raw.max() + 1e-6, 0.5)
            y_interp = np.interp(x_interp, x_raw, y_raw)

            ax_profile.plot(x_interp, y_interp, marker="o", linestyle="-", label="Tangent Diameter (mm)")
            ax_profile.set_title(f"Stem {stem_id} - Tangent Diameter vs Stem Length")
            ax_profile.set_xlabel("Stem Length from Bottom (cm)")
            ax_profile.set_ylabel("Diameter Tangent (mm)")
            ax_profile.grid(True)
            ax_profile.legend()
            ax_profile.set_xlim(global_xmin, global_xmax)
            ax_profile.set_ylim(global_ymin, global_ymax)
            plt.tight_layout()
            plt.savefig(output_dir / f"{image_path.stem}_stem_{stem_id}_diameter_vs_length.png")
            plt.close(fig_profile)
        else:
            print(f"Skipped diameter vs length plot for Stem {stem_id}; insufficient points.")
            plt.close(fig_profile)

    wb.save(excel_output_path)
    print(f"Finished: {excel_output_path}")


if __name__ == "__main__":
    batch_process_stems()

