"""
Horizontal scanline diameter measurement for maize/corn stem images.

This script measures stem diameter using a horizontal scanline approach. At each
node or internode sampling position, the point is first snapped to the skeleton.
A horizontal line is then scanned left and right through the filled stem mask to
find the two stem boundaries. The diameter is reported as the horizontal distance
between those boundaries.

Outputs
-------
1. Overlay image with node labels and horizontal diameter lines.
2. Excel file containing only horizontal scanline diameter measurements.

Expected folder structure
-------------------------
base_batch_dir/
    batch_*/
        piscart_whitebg/tif/*.tif
        yolo_output/labels/*.txt
        analysis_output/

The YOLO label file is expected to contain normalized node detections.
"""

import os
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from skimage import measure, morphology
import matplotlib.pyplot as plt


# -----------------------------------------------------------------------------
# User settings
# -----------------------------------------------------------------------------

PX_PER_MM_VERTICAL = 8.117542
PX_PER_MM_DIAMETER = 8.0

MIN_NODE_GAP_CM = 2.0
MIN_VALID_ROWS_FOR_NODE_SNAP = 3
SCANLINE_MEDIAN_HALF_WINDOW = 4

# Five representative diameter positions sampled within each internode.
INTERNODE_SAMPLE_FRACTIONS = [0.15, 0.35, 0.50, 0.65, 0.85]

# Drawing settings for publication overlays.
SHOW_SKELETON = False
SHOW_CONTOUR = True

COLOR_CONTOUR = (0, 255, 0)          # green
COLOR_SKELETON = (255, 255, 0)       # cyan
COLOR_SCANLINE = (0, 215, 255)       # bright yellow-green
COLOR_NODE_FILL = (0, 180, 0)        # dark green
COLOR_NODE_OUTLINE = (120, 255, 120) # light green
COLOR_NODE_TEXT = (255, 255, 255)    # white
COLOR_DIAMETER_TEXT = (0, 0, 255)    # red

LINE_THICKNESS = 7
TEXT_SCALE = 2.3
TEXT_THICKNESS = 5
LABEL_X_OFFSET = 32

# Update this path for your local project.
BASE_BATCH_DIR = Path(r"C:/Users/sayee/OneDrive/Desktop/Corn_Diameter/test_images/output_sessions")
DIAMETER_LOG_PATH = BASE_BATCH_DIR / "horizontal_scanline_completed_batches.txt"


# -----------------------------------------------------------------------------
# Input/output helpers
# -----------------------------------------------------------------------------

def read_completed_batches(log_path):
    """Read the list of batches that were already processed."""
    if not log_path.exists():
        return set()

    with open(log_path, "r", encoding="utf-8") as log_file:
        return {line.strip() for line in log_file if line.strip()}


def load_grayscale_image(image_path):
    """Load an image as normalized 8-bit grayscale."""
    image = cv2.imread(str(image_path), cv2.IMREAD_UNCHANGED)
    if image is None:
        raise FileNotFoundError(f"Image not found: {image_path}")

    if image.ndim == 3:
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    return cv2.normalize(image, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)


def sharpen_with_unsharp_mask(image, radius=3, amount=1.5):
    """Enhance local contrast using a standard unsharp-mask operation."""
    blurred = cv2.GaussianBlur(image, (0, 0), radius)
    return cv2.addWeighted(image, 1 + amount, blurred, -amount, 0)


def load_yolo_node_centers(label_file, image_shape, distance_threshold=250):
    """
    Load YOLO node detections and return node center coordinates in pixels.

    Nearby duplicate node detections are grouped, and the detection with the
    largest bounding-box height is retained as the representative node.
    """
    image_height, image_width = image_shape
    node_centers = []
    node_scores = []

    if not Path(label_file).exists():
        return np.zeros((0, 2), dtype=int)

    with open(label_file, "r", encoding="utf-8") as file:
        for line in file:
            parts = line.split()
            if len(parts) < 5:
                continue

            _, x_rel, y_rel, _, h_rel = map(float, parts[:5])
            node_centers.append((int(x_rel * image_width), int(y_rel * image_height)))
            node_scores.append(h_rel)

    if not node_centers:
        return np.zeros((0, 2), dtype=int)

    node_centers = np.array(node_centers, dtype=int)
    node_scores = np.array(node_scores, dtype=float)

    filtered_nodes = []
    used_indices = set()

    for i, point in enumerate(node_centers):
        if i in used_indices:
            continue

        cluster = [i]
        for j, other_point in enumerate(node_centers):
            if j == i or j in used_indices:
                continue
            if np.linalg.norm(point - other_point) < distance_threshold:
                cluster.append(j)

        best_index = max(cluster, key=lambda idx: node_scores[idx])
        filtered_nodes.append(tuple(node_centers[best_index]))
        used_indices.update(cluster)

    return np.array(filtered_nodes, dtype=int)


# -----------------------------------------------------------------------------
# Stem segmentation and skeleton utilities
# -----------------------------------------------------------------------------

def segment_stem_foreground(gray_image):
    """
    Segment stem pixels from the background using Otsu and adaptive thresholding.

    The threshold direction is selected automatically based on whether the image
    has a light or dark background.
    """
    sharpened = sharpen_with_unsharp_mask(gray_image)
    smoothed = cv2.bilateralFilter(sharpened, 7, 50, 50)

    is_light_background = np.mean(smoothed) > 128
    threshold_type = cv2.THRESH_BINARY_INV if is_light_background else cv2.THRESH_BINARY

    _, otsu_mask = cv2.threshold(smoothed, 0, 255, threshold_type + cv2.THRESH_OTSU)
    adaptive_mask = cv2.adaptiveThreshold(
        smoothed,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        threshold_type,
        31,
        2,
    )

    combined_mask = cv2.bitwise_or(otsu_mask, adaptive_mask)
    filled_mask = connect_and_fill_stem_mask(combined_mask, connect_distance=15)

    return sharpened, filled_mask


def connect_and_fill_stem_mask(binary_image, connect_distance=15):
    """Connect small gaps and fill holes inside the segmented stem mask."""
    binary = (binary_image > 0).astype(np.uint8) * 255

    kernel = cv2.getStructuringElement(
        cv2.MORPH_ELLIPSE,
        (connect_distance, connect_distance),
    )
    connected = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel, iterations=2)

    inverted = cv2.bitwise_not(connected)
    height, width = inverted.shape[:2]
    flood_mask = np.zeros((height + 2, width + 2), np.uint8)
    flood_filled = inverted.copy()
    cv2.floodFill(flood_filled, flood_mask, (0, 0), 255)

    holes = cv2.bitwise_not(flood_filled)
    filled = connected | holes
    filled = cv2.morphologyEx(filled, cv2.MORPH_CLOSE, np.ones((5, 5), np.uint8), iterations=1)

    return filled


def extract_stem_contours(filled_mask, min_area=500):
    """Extract external stem contours after segmentation."""
    contours, _ = cv2.findContours(filled_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = [contour for contour in contours if cv2.contourArea(contour) > min_area]
    return [cv2.approxPolyDP(contour, epsilon=1.0, closed=True) for contour in contours]


def skeletonize_stem_region(region_mask):
    """Create a one-pixel-wide centerline for one segmented stem region."""
    try:
        from cv2.ximgproc import thinning
        return thinning(region_mask, cv2.ximgproc.THINNING_ZHANGSUEN)
    except Exception:
        return morphology.skeletonize(region_mask > 0).astype(np.uint8) * 255


def snap_point_to_skeleton_row(center_x, center_y, skeleton_mask):
    """Snap a point to the nearest skeleton pixel on the same image row."""
    height, width = skeleton_mask.shape
    center_y = int(np.clip(center_y, 0, height - 1))
    center_x = int(np.clip(center_x, 0, width - 1))

    skeleton_x_values = np.where(skeleton_mask[center_y, :] > 0)[0]
    if len(skeleton_x_values) == 0:
        return center_x, center_y

    nearest_x = skeleton_x_values[np.argmin(np.abs(skeleton_x_values - center_x))]
    return int(nearest_x), int(center_y)


def snap_point_to_nearest_skeleton(center_x, center_y, skeleton_mask, search_radius=20):
    """Snap a point to the nearest skeleton pixel within a local search window."""
    height, width = skeleton_mask.shape
    center_x = int(np.clip(round(center_x), 0, width - 1))
    center_y = int(np.clip(round(center_y), 0, height - 1))

    y_min = max(0, center_y - search_radius)
    y_max = min(height, center_y + search_radius + 1)
    x_min = max(0, center_x - search_radius)
    x_max = min(width, center_x + search_radius + 1)

    patch = skeleton_mask[y_min:y_max, x_min:x_max]
    local_y, local_x = np.where(patch > 0)
    if len(local_x) == 0:
        return center_x, center_y

    skeleton_x = local_x + x_min
    skeleton_y = local_y + y_min
    squared_distance = (skeleton_x - center_x) ** 2 + (skeleton_y - center_y) ** 2
    nearest_index = int(np.argmin(squared_distance))

    return int(skeleton_x[nearest_index]), int(skeleton_y[nearest_index])


# -----------------------------------------------------------------------------
# Horizontal scanline diameter measurement
# -----------------------------------------------------------------------------

def find_horizontal_stem_boundaries(filled_region_mask, center_x, center_y):
    """
    Find left and right stem boundaries on the horizontal row through a point.

    The method uses the filled stem mask, not only the contour pixels. This makes
    the scanline robust to small gaps in the outline and returns the continuous
    foreground segment that contains, or is closest to, the supplied center point.
    """
    height, width = filled_region_mask.shape
    row_y = int(np.clip(round(center_y), 0, height - 1))
    center_x = int(np.clip(round(center_x), 0, width - 1))

    foreground_x = np.where(filled_region_mask[row_y, :] > 0)[0]
    if len(foreground_x) == 0:
        return None

    if filled_region_mask[row_y, center_x] == 0:
        center_x = int(foreground_x[np.argmin(np.abs(foreground_x - center_x))])

    left_x = center_x
    while left_x > 0 and filled_region_mask[row_y, left_x - 1] > 0:
        left_x -= 1

    right_x = center_x
    while right_x < width - 1 and filled_region_mask[row_y, right_x + 1] > 0:
        right_x += 1

    if right_x <= left_x:
        return None

    return {
        "left_boundary": (float(left_x), float(row_y)),
        "right_boundary": (float(right_x), float(row_y)),
        "center": (float(center_x), float(row_y)),
        "diameter_px": float(right_x - left_x),
    }


def measure_horizontal_scanline_diameter(filled_region_mask, skeleton_mask, center_x, center_y):
    """
    Measure diameter as the horizontal distance across the filled stem mask.

    Measurement steps:
    1. Snap the requested point to the nearest skeleton pixel.
    2. Use the skeleton row as the horizontal scanline location.
    3. Move left and right through foreground pixels until the stem boundary is reached.
    4. Report the distance between the two boundary points.
    """
    snapped_x, snapped_y = snap_point_to_nearest_skeleton(center_x, center_y, skeleton_mask, search_radius=20)
    measurement = find_horizontal_stem_boundaries(filled_region_mask, snapped_x, snapped_y)

    if measurement is None:
        return {
            "diameter_px": 0.0,
            "center": (float(snapped_x), float(snapped_y)),
            "pt1": (float(snapped_x), float(snapped_y)),
            "pt2": (float(snapped_x), float(snapped_y)),
        }

    return {
        "diameter_px": measurement["diameter_px"],
        "center": measurement["center"],
        "pt1": measurement["left_boundary"],
        "pt2": measurement["right_boundary"],
    }


def measure_robust_horizontal_scanline_diameter(
    filled_region_mask,
    skeleton_mask,
    center_x,
    center_y,
    half_window=SCANLINE_MEDIAN_HALF_WINDOW,
):
    """
    Return a stable horizontal diameter using nearby scanlines.

    A single image row may be affected by a small segmentation irregularity. To
    reduce this effect, diameters are measured across several neighboring rows,
    and the scanline closest to the median diameter is selected.
    """
    height, _ = filled_region_mask.shape
    center_x, center_y = snap_point_to_nearest_skeleton(center_x, center_y, skeleton_mask, search_radius=20)

    candidate_measurements = []
    for row_offset in range(-half_window, half_window + 1):
        row_y = int(np.clip(center_y + row_offset, 0, height - 1))
        row_x, row_y = snap_point_to_skeleton_row(center_x, row_y, skeleton_mask)
        measurement = measure_horizontal_scanline_diameter(
            filled_region_mask,
            skeleton_mask,
            row_x,
            row_y,
        )
        if measurement["diameter_px"] > 0:
            candidate_measurements.append(measurement)

    if not candidate_measurements:
        return measure_horizontal_scanline_diameter(
            filled_region_mask,
            skeleton_mask,
            center_x,
            center_y,
        )

    diameters = np.array([m["diameter_px"] for m in candidate_measurements], dtype=np.float32)
    median_diameter = float(np.median(diameters))
    selected_index = int(np.argmin(np.abs(diameters - median_diameter)))

    return candidate_measurements[selected_index]


# -----------------------------------------------------------------------------
# Drawing helpers
# -----------------------------------------------------------------------------

def draw_node_label(overlay_image, node_xy, node_label):
    """Draw a labeled node marker on the final overlay."""
    node_x, node_y = int(round(node_xy[0])), int(round(node_xy[1]))

    cv2.circle(overlay_image, (node_x, node_y), 34, COLOR_NODE_FILL, -1, lineType=cv2.LINE_AA)
    cv2.circle(overlay_image, (node_x, node_y), 34, COLOR_NODE_OUTLINE, 4, lineType=cv2.LINE_AA)

    cv2.putText(
        overlay_image,
        node_label,
        (node_x - 28, node_y + 12),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.15,
        COLOR_NODE_TEXT,
        3,
        lineType=cv2.LINE_AA,
    )


def draw_horizontal_diameter(overlay_image, scanline_measurement, diameter_mm, text_label):
    """Draw one horizontal scanline diameter and its numeric value."""
    x1, y1 = int(round(scanline_measurement["pt1"][0])), int(round(scanline_measurement["pt1"][1]))
    x2, y2 = int(round(scanline_measurement["pt2"][0])), int(round(scanline_measurement["pt2"][1]))

    cv2.line(
        overlay_image,
        (x1, y1),
        (x2, y2),
        COLOR_SCANLINE,
        LINE_THICKNESS,
        lineType=cv2.LINE_AA,
    )

    text_x = int(round(max(x1, x2))) + LABEL_X_OFFSET
    text_y = int(round((y1 + y2) / 2))
    label_text = f"{text_label}: {diameter_mm:.1f}"

    cv2.putText(
        overlay_image,
        label_text,
        (text_x, text_y + 5),
        cv2.FONT_HERSHEY_SIMPLEX,
        TEXT_SCALE,
        COLOR_DIAMETER_TEXT,
        TEXT_THICKNESS,
        lineType=cv2.LINE_AA,
    )


# -----------------------------------------------------------------------------
# Batch processing
# -----------------------------------------------------------------------------

def batch_process_stems():
    """Process all batch folders and save scanline-only outputs."""
    completed_batches = read_completed_batches(DIAMETER_LOG_PATH)

    for batch_folder in sorted(BASE_BATCH_DIR.glob("batch_*")):
        batch_name = batch_folder.name

        if batch_name in completed_batches:
            print(f"Skipping completed batch: {batch_name}")
            continue

        input_dir = batch_folder / "piscart_whitebg" / "tif"
        labels_dir = batch_folder / "yolo_output" / "labels"
        output_dir = batch_folder / "analysis_output"
        output_dir.mkdir(parents=True, exist_ok=True)

        tif_images = sorted(input_dir.glob("*.tif"))
        if not tif_images:
            print(f"No TIF images found in {batch_name}; skipping.")
            continue

        processed_count = 0

        for image_path in tif_images:
            image_stem = image_path.stem
            base_name = image_stem.replace("_whitebg", "").replace("_blackbg", "")
            label_path = labels_dir / f"{base_name}_blackbg.txt"
            image_output_dir = output_dir / base_name
            image_output_dir.mkdir(parents=True, exist_ok=True)

            print(f"Processing {base_name} in batch {batch_name}")
            try:
                process_single_stem_image(image_path, label_path, image_output_dir)
                processed_count += 1
            except Exception as error:
                print(f"Error processing {base_name}: {error}")

        if processed_count > 0:
            with open(DIAMETER_LOG_PATH, "a", encoding="utf-8") as log_file:
                log_file.write(batch_name + "\n")
            print(f"Batch {batch_name} completed and logged ({processed_count} images).")
        else:
            print(f"Batch {batch_name} had no valid images; not logged.")


def process_single_stem_image(image_path, label_path, output_dir, global_node_start=1):
    """Process one image and save horizontal scanline measurements."""
    gray_image = load_grayscale_image(image_path)
    display_image, filled_mask = segment_stem_foreground(gray_image)
    contours = extract_stem_contours(filled_mask)

    overlay_image = cv2.cvtColor(display_image.copy(), cv2.COLOR_GRAY2BGR)
    if SHOW_CONTOUR:
        cv2.drawContours(overlay_image, contours, -1, COLOR_CONTOUR, thickness=2, lineType=cv2.LINE_AA)

    detected_nodes = load_yolo_node_centers(label_path, gray_image.shape)

    # Separate touching or nearby stems into connected regions.
    merged_mask = cv2.morphologyEx(filled_mask, cv2.MORPH_CLOSE, np.ones((15, 3), np.uint8), iterations=2)
    labeled_mask = measure.label(merged_mask)
    raw_regions = measure.regionprops(labeled_mask)

    def bottom_x_position(region):
        region_y, region_x = np.nonzero(labeled_mask == region.label)
        if len(region_y) == 0:
            return 1e18
        bottom_y = region_y.max()
        x_at_bottom = region_x[region_y == bottom_y]
        if len(x_at_bottom) == 0:
            return float(np.median(region_x))
        return float(np.median(x_at_bottom))

    stem_regions = sorted(raw_regions, key=bottom_x_position)

    measurement_rows = []
    global_node_index = int(global_node_start)
    stem_counter = 0

    for region in stem_regions:
        if region.area < 5000:
            continue

        region_mask = (labeled_mask == region.label).astype(np.uint8)
        region_mask = morphology.binary_closing(region_mask, morphology.disk(3))
        region_mask = morphology.binary_dilation(region_mask, morphology.disk(1))
        region_mask = region_mask.astype(np.uint8)

        skeleton = skeletonize_stem_region((region_mask * 255).astype(np.uint8))
        if len(np.nonzero(skeleton)[0]) < 10:
            continue

        if SHOW_SKELETON:
            skeleton_y, skeleton_x = np.where(skeleton > 0)
            for x_coord, y_coord in zip(skeleton_x, skeleton_y):
                overlay_image[y_coord, x_coord] = COLOR_SKELETON

        stem_y, stem_x = np.nonzero(region_mask)
        if len(stem_y) == 0:
            continue

        stem_counter += 1
        stem_id = stem_counter

        stem_bottom_y = int(stem_y.max())
        stem_top_y = int(stem_y.min())
        stem_total_length_cm = (stem_bottom_y - stem_top_y) / PX_PER_MM_VERTICAL / 10.0

        # Add pseudo-nodes near the base and top when needed so that the full
        # visible stem is represented in the output table.
        actual_bottom_y = int(stem_y.max())
        actual_bottom_x = int(np.median(stem_x[stem_y == actual_bottom_y]))

        offset_bottom_y = max(0, actual_bottom_y - 5)
        if np.any(stem_y == offset_bottom_y):
            offset_bottom_x = int(np.median(stem_x[stem_y == offset_bottom_y]))
        else:
            offset_bottom_x = actual_bottom_x

        bottom_actual = measure_robust_horizontal_scanline_diameter(
            region_mask,
            skeleton,
            actual_bottom_x,
            actual_bottom_y,
        )
        bottom_offset = measure_robust_horizontal_scanline_diameter(
            region_mask,
            skeleton,
            offset_bottom_x,
            offset_bottom_y,
        )

        if bottom_offset["diameter_px"] > bottom_actual["diameter_px"]:
            bottom_node = (offset_bottom_x, offset_bottom_y)
        else:
            bottom_node = (actual_bottom_x, actual_bottom_y)

        top_node = (int(np.median(stem_x[stem_y == stem_top_y])), stem_top_y)

        one_mm_y_shift = int(round(PX_PER_MM_VERTICAL))
        bottom_node = (bottom_node[0], bottom_node[1] - one_mm_y_shift)
        top_node = (top_node[0], top_node[1] + one_mm_y_shift)

        min_row, min_col, max_row, max_col = region.bbox

        def point_inside_region(point):
            return min_col <= point[0] <= max_col and min_row <= point[1] <= max_row

        region_nodes = [point for point in detected_nodes if point_inside_region(point)]
        region_nodes.append(bottom_node)

        top_threshold_y = min_row + (max_row - min_row) * 0.15
        top_node_already_detected = any(
            point[1] <= top_threshold_y for point in detected_nodes if point_inside_region(point)
        )
        if not top_node_already_detected:
            region_nodes.append(top_node)

        region_nodes = sorted(region_nodes, key=lambda point: point[1], reverse=True)

        # Merge node detections that are too close to represent real internodes.
        collapsed_nodes = []
        node_index = 0
        while node_index < len(region_nodes):
            if node_index == len(region_nodes) - 1:
                collapsed_nodes.append(region_nodes[node_index])
                break

            lower_node = region_nodes[node_index]
            upper_node = region_nodes[node_index + 1]
            gap_cm = abs(lower_node[1] - upper_node[1]) / PX_PER_MM_VERTICAL / 10.0

            if gap_cm < MIN_NODE_GAP_CM:
                candidate_fractions = [0.0, 1 / 6, 2 / 6, 3 / 6, 4 / 6, 5 / 6, 1.0]
                best_point = lower_node
                largest_diameter_px = -1.0

                for fraction in candidate_fractions:
                    candidate_x = int((1 - fraction) * lower_node[0] + fraction * upper_node[0])
                    candidate_y = int((1 - fraction) * lower_node[1] + fraction * upper_node[1])
                    candidate_x, candidate_y = snap_point_to_nearest_skeleton(
                        candidate_x,
                        candidate_y,
                        skeleton,
                        search_radius=20,
                    )
                    scanline = measure_robust_horizontal_scanline_diameter(
                        region_mask,
                        skeleton,
                        candidate_x,
                        candidate_y,
                    )
                    if scanline["diameter_px"] > largest_diameter_px:
                        largest_diameter_px = scanline["diameter_px"]
                        best_point = (candidate_x, candidate_y)

                collapsed_nodes.append(best_point)
                node_index += 2
            else:
                collapsed_nodes.append(lower_node)
                node_index += 1

        region_nodes = [
            snap_point_to_nearest_skeleton(x_coord, y_coord, skeleton, search_radius=20)
            for x_coord, y_coord in collapsed_nodes
        ]

        # Refine internal node locations by searching nearby rows for the widest
        # horizontal scanline. This places node points near local swelling regions.
        if len(region_nodes) >= 3:
            refined_nodes = list(region_nodes)
            for k in range(1, len(refined_nodes) - 1):
                previous_y = refined_nodes[k - 1][1]
                current_x, current_y = refined_nodes[k]
                next_y = refined_nodes[k + 1][1]

                search_y1 = int((previous_y + current_y) / 2)
                search_y2 = int((current_y + next_y) / 2)
                search_y1 = int(np.clip(search_y1, 0, region_mask.shape[0] - 1))
                search_y2 = int(np.clip(search_y2, 0, region_mask.shape[0] - 1))

                best_candidate = None
                valid_rows = 0
                for row_y in range(min(search_y1, search_y2), max(search_y1, search_y2) + 1):
                    row_x, row_y = snap_point_to_skeleton_row(current_x, row_y, skeleton)
                    scanline = measure_robust_horizontal_scanline_diameter(region_mask, skeleton, row_x, row_y)
                    diameter_px = scanline["diameter_px"]
                    if diameter_px <= 0:
                        continue
                    valid_rows += 1
                    if best_candidate is None or diameter_px > best_candidate[0]:
                        best_candidate = (diameter_px, row_x, row_y)

                if best_candidate is not None and valid_rows >= MIN_VALID_ROWS_FOR_NODE_SNAP:
                    refined_nodes[k] = (int(best_candidate[1]), int(best_candidate[2]))

            region_nodes = sorted(refined_nodes, key=lambda point: point[1], reverse=True)

        if len(region_nodes) < 2:
            continue

        # Assign node names from the base upward.
        node_names = []
        if stem_id == 1:
            node_names.append(f"node_{global_node_index}.0")
        else:
            node_names.append(f"node_{global_node_index}.1")

        for _ in range(1, len(region_nodes)):
            global_node_index += 1
            node_names.append(f"node_{global_node_index}.0")

        numeric_node_labels = []
        for node_name in node_names:
            node_number = node_name.split("_")[1].split(".")[0]
            numeric_node_labels.append(f"N{node_number}")

        for node_label, node_xy in zip(numeric_node_labels, region_nodes):
            draw_node_label(overlay_image, node_xy, node_label)

        def add_node_measurement_row(node_position, node_name):
            """Add one node row using the horizontal scanline diameter method."""
            node_x, node_y = node_position
            point_length_cm = (stem_bottom_y - node_y) / PX_PER_MM_VERTICAL / 10.0
            point_length_cm = max(0.0, point_length_cm)

            if node_name == node_names[0]:
                stem_length_cm = 0.0
            elif node_name == node_names[-1]:
                stem_length_cm = stem_total_length_cm
            else:
                stem_length_cm = point_length_cm

            scanline = measure_robust_horizontal_scanline_diameter(region_mask, skeleton, node_x, node_y)
            diameter_mm = scanline["diameter_px"] / PX_PER_MM_DIAMETER if scanline["diameter_px"] > 0 else 0.0

            measurement_rows.append({
                "Stem": stem_id,
                "Type": node_name,
                "Y(px)": int(node_y),
                "X(px)": int(node_x),
                "Point Length (cm)": round(point_length_cm, 2),
                "Stem Length (cm)": round(stem_length_cm, 2),
                "Diameter Horizontal Scanline (mm)": round(diameter_mm, 2),
            })

        # Write rows in physical order along the stem:
        # node -> D1..D5 -> next node. This order is required for correct
        # internode length calculation and clean Excel summaries.
        for i in range(len(region_nodes) - 1):
            lower_node = region_nodes[i]
            upper_node = region_nodes[i + 1]
            lower_name = node_names[i]
            upper_name = node_names[i + 1]

            if i == 0:
                add_node_measurement_row(lower_node, lower_name)

            for sample_index, fraction in enumerate(INTERNODE_SAMPLE_FRACTIONS, start=1):
                sample_x = int((1 - fraction) * lower_node[0] + fraction * upper_node[0])
                sample_y = int((1 - fraction) * lower_node[1] + fraction * upper_node[1])
                sample_x, sample_y = snap_point_to_nearest_skeleton(sample_x, sample_y, skeleton, search_radius=20)

                scanline = measure_robust_horizontal_scanline_diameter(
                    region_mask,
                    skeleton,
                    sample_x,
                    sample_y,
                )
                diameter_mm = scanline["diameter_px"] / PX_PER_MM_DIAMETER if scanline["diameter_px"] > 0 else 0.0

                center_x, center_y = scanline["center"]
                center_x = int(round(center_x))
                center_y = int(round(center_y))
                point_length_cm = (stem_bottom_y - center_y) / PX_PER_MM_VERTICAL / 10.0
                point_length_cm = max(0.0, point_length_cm)

                sample_label = f"D{sample_index}"
                measurement_rows.append({
                    "Stem": stem_id,
                    "Type": f"{sample_label}_between_{lower_name}_{upper_name}",
                    "Y(px)": center_y,
                    "X(px)": center_x,
                    "Point Length (cm)": round(point_length_cm, 2),
                    "Stem Length (cm)": round(point_length_cm, 2),
                    "Diameter Horizontal Scanline (mm)": round(diameter_mm, 2),
                })

                draw_horizontal_diameter(overlay_image, scanline, diameter_mm, sample_label)

            add_node_measurement_row(upper_node, upper_name)

    overlay_path = output_dir / f"{image_path.stem}_horizontal_scanline_overlay_5points.png"
    cv2.imwrite(str(overlay_path), overlay_image)
    print(f"Overlay saved: {overlay_path}")

    save_measurements_to_excel(measurement_rows, output_dir, image_path.stem)


def save_measurements_to_excel(measurement_rows, output_dir, image_stem):
    """Save scanline measurements with grouped sheets, summaries, and plots."""
    data_frame = pd.DataFrame(measurement_rows)
    if data_frame.empty:
        print(f"No valid horizontal scanline measurements found for {image_stem}.")
        return

    # Internode length is calculated only on node rows. Because rows are written
    # in the order node -> D points -> next node, this produces the true
    # node-to-node distance for each internode.
    internode_lengths = []
    last_node_length = None
    current_stem = None

    for _, row in data_frame.iterrows():
        row_type = str(row["Type"])
        stem_id = row["Stem"]

        if stem_id != current_stem:
            last_node_length = None
            current_stem = stem_id

        if row_type.lower().startswith("node_"):
            current_length = float(row["Point Length (cm)"])
            if last_node_length is None:
                internode_lengths.append("")
            else:
                internode_lengths.append(round(current_length - last_node_length, 2))
            last_node_length = current_length
        else:
            internode_lengths.append("")

    data_frame["Internode Length (cm)"] = internode_lengths

    ordered_columns = [
        "Stem",
        "Type",
        "Y(px)",
        "X(px)",
        "Point Length (cm)",
        "Stem Length (cm)",
        "Internode Length (cm)",
        "Diameter Horizontal Scanline (mm)",
    ]
    data_frame = data_frame[ordered_columns]

    data_frame["Stem Length (cm)"] = pd.to_numeric(data_frame["Stem Length (cm)"], errors="coerce")
    data_frame["Diameter Horizontal Scanline (mm)"] = pd.to_numeric(
        data_frame["Diameter Horizontal Scanline (mm)"],
        errors="coerce",
    )

    global_xmin = 0.0
    global_xmax = float(data_frame["Stem Length (cm)"].max()) if data_frame["Stem Length (cm)"].notna().any() else 0.0

    diameter_values = data_frame["Diameter Horizontal Scanline (mm)"].dropna()
    diameter_values = diameter_values[(diameter_values > 0) & (diameter_values < 200)]
    global_ymin = float(diameter_values.min()) if len(diameter_values) else 0.0
    global_ymax = float(diameter_values.max()) if len(diameter_values) else 1.0
    pad = 0.05 * (global_ymax - global_ymin + 1e-6)
    global_ymin -= pad
    global_ymax += pad

    excel_path = output_dir / f"{image_stem}_horizontal_scanline_5point_measurements.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="All_Points", index=False)

    workbook = load_workbook(excel_path)
    all_points_sheet = workbook["All_Points"]
    current_row = all_points_sheet.max_row + 2

    for stem_id in sorted(data_frame["Stem"].unique()):
        stem_frame = data_frame[data_frame["Stem"] == stem_id].copy()
        if stem_frame.empty:
            continue

        header_cell = all_points_sheet.cell(row=current_row, column=2, value=f"Stem {stem_id} - Horizontal Scanline Analysis")
        header_cell.font = Font(bold=True, color="800080")

        total_length = float(stem_frame["Stem Length (cm)"].max())
        length_cell = all_points_sheet.cell(row=current_row, column=6, value=f"Total Length = {round(total_length, 2)}cms")
        length_cell.font = Font(bold=True)

        sheet_name = f"Stem_{stem_id}"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        header_cell.hyperlink = Hyperlink(ref="", location=f"'{sheet_name}'!A1", display=header_cell.value)

        for col_idx, col_name in enumerate(data_frame.columns, start=1):
            all_points_sheet.cell(row=current_row + 1, column=col_idx, value=col_name).font = Font(bold=True)

        data_start_row = current_row + 2
        for i, row_data in enumerate(stem_frame.itertuples(index=False)):
            for col_idx, value in enumerate(row_data, start=1):
                all_points_sheet.cell(row=data_start_row + i, column=col_idx, value=value)

        current_row = data_start_row + len(stem_frame) + 4

        stem_sheet = workbook[sheet_name]
        max_length = int(np.ceil(stem_frame["Stem Length (cm)"].max()))
        bins = list(range(0, max_length + 10, 10))
        if len(bins) < 2:
            bins = [0, 10]

        stem_summary_frame = stem_frame.copy()
        stem_summary_frame["Segment"] = pd.cut(
            stem_summary_frame["Stem Length (cm)"],
            bins=bins,
            right=False,
            include_lowest=True,
        ).astype(str)

        segment_order = stem_summary_frame["Segment"].dropna().unique()
        node_rows = stem_summary_frame[stem_summary_frame["Type"].str.contains("^node_", regex=True, case=False)]

        node_counts = node_rows.groupby("Segment", observed=False).size()
        internode_lengths = pd.to_numeric(node_rows["Internode Length (cm)"], errors="coerce")
        avg_internode_length = node_rows.assign(_internode_length=internode_lengths).groupby(
            "Segment", observed=False
        )["_internode_length"].mean()
        avg_diameter = stem_summary_frame.groupby("Segment", observed=False)["Diameter Horizontal Scanline (mm)"].mean()

        summary = pd.DataFrame({
            "Segment": segment_order.astype(str),
            "Node Count": node_counts.reindex(segment_order).fillna(0).astype(int).values,
            "Avg Internode Length (cm)": avg_internode_length.reindex(segment_order).values,
            "Avg Horizontal Diameter (mm)": avg_diameter.reindex(segment_order).values,
            "Curvature (deg)": [round(90 - i * 15.23, 2) for i in range(len(segment_order))],
        })

        for col_idx, header in enumerate(summary.columns, start=1):
            stem_sheet.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
        for row_idx, row_data in enumerate(summary.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                stem_sheet.cell(row=row_idx, column=col_idx, value=value)

        offset = len(summary) + 2
        stem_sheet.cell(row=offset, column=6, value=f"Total Length = {round(total_length, 2)}cms").font = Font(bold=True)

        for col_idx, col_name in enumerate(data_frame.columns, start=1):
            stem_sheet.cell(row=offset + 2, column=col_idx, value=col_name).font = Font(bold=True)
        for i, row_data in enumerate(stem_summary_frame.itertuples(index=False)):
            for col_idx, value in enumerate(row_data, start=1):
                stem_sheet.cell(row=offset + 3 + i, column=col_idx, value=value)

        # Segment-wise structural plot.
        summary_plot = summary.copy()
        summary_plot["Segment Mid"] = summary_plot["Segment"].str.extract(r"(\d+)").astype(float)
        summary_plot = summary_plot.sort_values("Segment Mid")

        fig, ax = plt.subplots(figsize=(12, 6))
        x_labels = summary_plot["Segment"].tolist()
        x_pos = np.arange(len(x_labels))
        ax.plot(x_pos, summary_plot["Node Count"], marker="o", label="Node Count")
        ax.plot(x_pos, summary_plot["Avg Horizontal Diameter (mm)"], marker="s", label="Avg Horizontal Diameter (mm)")
        ax.plot(x_pos, summary_plot["Avg Internode Length (cm)"], marker="^", label="Avg Internode Length (cm)")
        ax.plot(x_pos, summary_plot["Curvature (deg)"] / 10, marker="d", label="Curvature (deg / 10)")
        ax.set_title(f"Stem {stem_id} - Segment-wise Horizontal Scanline Analysis")
        ax.set_xlabel("Segment (cm)")
        ax.set_ylabel("Value")
        ax.set_xticks(x_pos)
        ax.set_xticklabels(x_labels, rotation=45, ha="right")
        ax.legend()
        ax.grid(True)
        plt.tight_layout()
        plt.savefig(output_dir / f"{image_stem}_stem_{stem_id}_segments.png")
        plt.close(fig)

        # Diameter profile along stem axis.
        fig_profile, ax_profile = plt.subplots(figsize=(10, 5))
        raw_curve = stem_frame[["Stem Length (cm)", "Diameter Horizontal Scanline (mm)"]].copy()
        raw_curve = raw_curve.dropna().sort_values("Stem Length (cm)")
        raw_curve = raw_curve.drop_duplicates(subset=["Stem Length (cm)"], keep="first")

        if raw_curve.shape[0] >= 2:
            x_raw = raw_curve["Stem Length (cm)"].values
            y_raw = raw_curve["Diameter Horizontal Scanline (mm)"].values
            x_interp = np.arange(x_raw.min(), x_raw.max() + 1e-6, 0.5)
            y_interp = np.interp(x_interp, x_raw, y_raw)

            ax_profile.plot(x_interp, y_interp, marker="o", linestyle="-", label="Horizontal Diameter (mm)")
            ax_profile.set_title(f"Stem {stem_id} - Horizontal Scanline Diameter vs Stem Length")
            ax_profile.set_xlabel("Stem Length from Bottom (cm)")
            ax_profile.set_ylabel("Diameter Horizontal Scanline (mm)")
            ax_profile.grid(True)
            ax_profile.legend()
            ax_profile.set_xlim(global_xmin, global_xmax)
            ax_profile.set_ylim(global_ymin, global_ymax)
            plt.tight_layout()
            plt.savefig(output_dir / f"{image_stem}_stem_{stem_id}_diameter_vs_length.png")
            plt.close(fig_profile)
        else:
            print(f"Skipped diameter vs length plot for Stem {stem_id}; insufficient points.")
            plt.close(fig_profile)

    workbook.save(excel_path)
    print(f"Measurements saved: {excel_path}")


if __name__ == "__main__":
    batch_process_stems()
